import tkinter as tk
from tkinter import filedialog

import pypinyin
from openpyxl import load_workbook, Workbook
from pypinyin import pinyin

import copy
import json
import os

Dm_Code_List = []
Config = {
    "row_num": 4,
    "unit": "得分",
    "level": "市级",
    "data_time": "2023-06-23 1:00:00",
    "flag": None
}


# 汉字转拼音
def to_pinyin(word):
    s = ''
    for w in pinyin(word, style=pypinyin.NORMAL):
        s += ''.join(w)
    return s


# 判断letter_range字母集合是否为target_range字母集合的自己
# C:D ⊆ C:I
def is_within_range(letter_range, target_range):
    # 将字母区间转换为ASCII码值区间
    start = ord(letter_range[0])
    end = ord(letter_range[-1])
    letter_range_ascii = range(start, end + 1)

    # 将目标区间转换为ASCII码值区间
    start = ord(target_range[0])
    end = ord(target_range[-1])
    target_range_ascii = range(start, end + 1)

    # 判断字母区间是否在目标区间中
    return set(letter_range_ascii).issubset(set(target_range_ascii))


# 生成新行
def generate_new_rows(row, islast, old_names):
    names = []
    name = None

    for cell in row:
        # 值不为空，当前单元格内容作为 title
        if cell.value is not None:
            name = cell.value
            start = cell.coordinate[0]
            end = cell.coordinate[0]
            # 根据当前单元格区间，在 old_names 中查找符合子集条件的对象
            # 然后把该对象的  name 赋值给当前单元格的 parent 属性
            parentObj = next((n for n in old_names if is_within_range(f"{start}:{end}", f"{n['start']}:{n['end']}")), None)
            parent = '##'
            if parentObj is not None:
                parent = parentObj['name']
            names.append({'name': name, 'name_en': to_pinyin(name), 'parent': parent, 'start': start, 'end': end})

        # 值为空，更新合并单元格的最后一个列，最后一行表头跳过 end 更新
        if islast is False:
            if cell.value is None:
                end = cell.coordinate[0]
                for t in names:
                    if t['name'] == name:
                        t['end'] = end

    return names


# 单元格区间求和
def cell_range_sum(cell_range):
    _sum = 0
    for cell in cell_range:
        if cell.value is None:
            break
        if isinstance(cell.value, str):
            _sum = cell.value
        else:
            if cell.data_type == "f":  # 判断单元格是否为公式
                cell.value = cell.calculate_value()
            _sum += cell.value
    return _sum


# 根据关键字查看部门，并返回 dm_code
def find_dm_code(keyword):
    dmMap = next((dm for dm in Dm_Code_List if dm['jc'].find(keyword) != -1), None)
    return dmMap['dm']


def parse_excel(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    sheet = wb.active
    new_table_data = []
    column_data = []
    rowIndex = 1

    # 对数据进行某种操作
    for row in sheet.iter_rows():
        # 获取所有表头，以及表头所处的单元格『合并单元格』范围
        if rowIndex <= Config['row_num']:
            column_data += generate_new_rows(row, islast=rowIndex == Config['row_num'], old_names=column_data)

        else:
            # 缓存当前行的组织编码
            code = ''
            row_data = []
            for col in column_data:
                range_key = f"{col['start']}{rowIndex}:{col['end']}{rowIndex}"
                cell_range = sheet[range_key][0]
                col['value'] = cell_range_sum(cell_range)
                # 转换部门 code
                if isinstance(col['value'], str):
                    code = find_dm_code(col['value'])
                    col['dm_code'] = find_dm_code(col['value'])
                    print(f'组织机构: {col}')
                    print('')
                else:
                    col['dm_code'] = code
                    col['适用部门编号'] = col['dm_code']
                    col['指标中文名称'] = col['name']
                    col['指标英文名称'] = col['name_en']
                    col['指标值'] = col['value']
                    col['父指标名称'] = col['parent']
                    col['指标单位'] = Config['unit']
                    col['百分比标志位'] = Config['flag']
                    col['适配层级'] = Config['level']
                    col['业务时间'] = Config['data_time']
                # 对字典对象进行深拷贝，不然会出现下一行数据覆盖之前行数据的情况
                row_data.append(copy.deepcopy(col))
            # 移除第一条组织信息
            row_data.pop(0)
            new_table_data.append(row_data)
        rowIndex += 1

        # for cell in row:
        #     # 输出每个属性的值
        #     for attr in dir(cell):
        #         if attr[0] != '_':
        #             print(f"cell {attr} = {getattr(cell, attr)}")
        #     print("")

    return new_table_data


def main():
    # 获取当前程序运行的目录
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建dm_code文件的完整路径
    dmcode_file = os.path.join(current_dir, "dmcode.json")

    # 加载dmcode文件
    with open(dmcode_file, "r") as file:
        dmcode = json.load(file)
    global Dm_Code_List
    Dm_Code_List = dmcode
    print(f'dmcode = {dmcode}')

    # 构建config文件的完整路径
    config_file = os.path.join(current_dir, "config.json")

    # 加载config文件
    with open(config_file, "r") as file:
        config = json.load(file)
    global Config
    Config = config
    print(f'Config = {config}')

    # 打开文件选择对话框，选择Excel文件
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    table_data = parse_excel(file_path)

    print(f'new_table_data = {table_data}')

    # 创建一个新的工作簿
    new_wb = Workbook()

    # 获取默认的工作表
    ws = new_wb.active

    # 定义表头
    headers = ["适用部门编号", "指标中文名称", "指标英文名称", "指标值", "指标单位", "百分比标志位", "适配层级", '父指标名称', '业务时间']

    # 写入表头
    ws.append(headers)

    # 遍历二维数组中的每个对象
    for row in table_data:
        for item in row:
            # 创建一个新行
            new_row = []

            # 将每个属性的值添加到新行中
            for header in headers:
                new_row.append(item[header])

            # 将新行添加到工作表中
            ws.append(new_row)

    # 保存新的Excel文件
    new_file_path = filedialog.asksaveasfilename(defaultextension="output.xlsx")
    new_wb.save(new_file_path)


# 创建UI窗口
root = tk.Tk()
root.title("Excel解析程序")

# 创建文件选择按钮
button = tk.Button(root, text="选择Excel文件", command=main)
button.pack(padx=20, pady=20)

# 运行UI主循环
root.mainloop()
